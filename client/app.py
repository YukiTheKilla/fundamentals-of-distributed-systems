from flask import Flask, render_template, request, redirect, url_for, session
import socket
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter


app = Flask(__name__)
app.secret_key = 'kalisto'

user_data_file = './client/user_data.json'

try:
    with open(user_data_file, 'r') as file:
        users = json.load(file)
except FileNotFoundError:
    users = []

def welcome_user(username):
    return f"Welcome, {username}! You are now logged in."

def receive_file(client_socket, file_name):
    # Receive the file data from the server
    file_data = client_socket.recv(1024)

    # Specify the path to save the file
    file_path = f'./client/downloads/{file_name}'

    with open(file_path, 'wb') as file:
        while file_data:
            file.write(file_data)
            file_data = client_socket.recv(1024)

    print(f"File received and saved at: {file_path}")
    
def receive_json_data(client_socket):
    # Receive the JSON data from the server
    received_data = b''
    while True:
        data_chunk = client_socket.recv(1024)
        if not data_chunk:
            break
        received_data += data_chunk

    json_data = received_data.decode('utf-8')
    return json_data

def get_json_file_list():
    # Connect to the server
    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    client.connect(('127.0.0.1', 2222))

    try:
        # Receive the JSON data from the server
        data = client.recv(1024).decode()

        # Parse the JSON data and print it
        json_files = json.loads(data)
        return json_files

    finally:
        # Close the client socket
        client.close()

@app.route('/')
def main():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Basic validation (replace with more robust validation)
        if not username or not password:
            return "Please provide both username and password."

        # Check if the username is already taken
        if any(user['username'] == username for user in users):
            return "Username already taken. Choose a different one."

        # Store user data
        users.append({'username': username, 'password': password})

        # Save user data to the JSON file
        with open(user_data_file, 'w') as file:
            json.dump(users, file, indent=2)

        return f"Registration successful! Welcome, {username}!"

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Basic validation 
        if not username or not password:
            return "Please provide both username and password."

        # Check if the user exists and the password is correct
        user = next((user for user in users if user['username'] == username and user['password'] == password), None)

        if user:
            # Store the username in the session
            session['username'] = username

            # Redirect to the route that calls the get_json_file_list function
            return redirect(url_for('get_json_files'))
        else:
            return "Invalid username or password."

    return render_template('login.html')

@app.route('/get_json_files')
def get_json_files():
    if 'username' in session:
        # Call the get_json_file_list function and pass the result to the template
        json_files = get_json_file_list()
        return render_template('json_files.html', json_files=json_files)
    else:
        return redirect(url_for('login'))

@app.route('/send_json_request', methods=['POST'])
def send_request():
    if 'username' in session:
        selected_file = request.form.get('selected_file')

        # Connect to the server on port 6666
        client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        client.connect(('127.0.0.1', 3333))

        try:
            # Send the selected file name to the server
            client.send(selected_file.encode('utf-8'))

            # Receive the JSON data from the server
            json_data = receive_json_data(client) 

            # Parse JSON data
            data = json.loads(json_data)

            # Create a new Workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write headers to the sheet
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = header

            # Write data to the sheet
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    sheet[f"{col_letter}{row_num}"] = row_data[header]

            # Specify the path to save the XLSX file
            xlsx_file_path = os.path.join('./client/downloads', f"{selected_file.replace('.json', '')}.xlsx")

            # Save Workbook to XLSX file
            workbook.save(xlsx_file_path)

            print(f"File received, converted, and saved at: {xlsx_file_path}")

            # Redirect back to the file selection page after download
            return redirect(url_for('get_json_files'))

        except json.decoder.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")

        finally:
            # Close the client socket
            client.close()

    else:
        return redirect(url_for('login'))
    
if __name__ == '__main__':
    while True:
        try:
            app.run(debug=True)
        except KeyboardInterrupt:
            print("Flask app interrupted. Exiting.")
            break